#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ActiveOrders → SALES_KSP_CRM_V3.xlsx appender (Excel-safe)

What this does
--------------
1) Reads every *.xlsx in --orders-dir that looks like Kaspi "ActiveOrders" export.
2) Normalizes Russian headers to a canonical set.
3) Filters rows:
   * Статус == "Ожидает передачи курьеру"    (config via --status)
   * Требуется подписание == "Не требуется"  (if such a column exists; config via --signature)
   * Плановая дата передачи курьеру <= --date-end (default: today, local)
4) Builds a staging block whose **columns match your CRM table slice**:
   from column header "Date" (we fill this separately) and
   from "OrderID/№ заказа" through "Склад передачи КД".
   Any CRM columns that have no counterpart in the export (e.g. MY_SIZE) are left blank.
5) Opens Excel invisibly via xlwings, **adds N rows to your table** so formulas propagate,
   and pastes:
      - today's date into the Date column for the N new rows
      - the staging block into [OrderID .. Склад передачи КД] for the N new rows
6) Saves & closes (no Excel UI).

Design notes
------------
- We never write the CRM file with openpyxl → avoids "Repair" alerts.
- No OrderID de-dup (per your rule).
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from dateutil import parser as dtp
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


OPS_ROOT = Path(__file__).resolve().parent
DEFAULT_ACTIVE_ORDERS_DIR = Path(os.environ.get("KASPI_ACTIVE_ORDERS_DIR", OPS_ROOT / "ActiveOrders"))
DEFAULT_CRM_WORKBOOK = Path(os.environ.get("KASPI_CRM_WORKBOOK", OPS_ROOT / "SALES_KSP_CRM_V3.xlsx"))
DEFAULT_CRM_SHEET = os.environ.get("KASPI_CRM_SHEET", "SALES_KSP_CRM_1")
DEFAULT_CRM_TABLE = os.environ.get("KASPI_CRM_TABLE", "CRM")
DEFAULT_KASPI_STATUS = os.environ.get("KASPI_STATUS", "Ожидает передачи курьеру")
DEFAULT_KASPI_SIGNATURE = os.environ.get("KASPI_SIGNATURE", "Не требуется")


# ---------- Helpers ----------

def norm(s: str) -> str:
    """lowercase, strip spaces/punct; cyrillic-friendly."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("ё", "е")
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s", "", s)
    return s

# Canonical header keys we care about
CANON = {
    "order_id": ["№заказа", "номерзаказа", "orderid", "заказ", "номерзаказа№"],
    "status": ["статус"],
    "signature": ["требуетсяподписание"],
    "handover": ["плановаядатапередачикурьеру", "плановаядатапередачи"],
    "offer_name": ["названиетоваравkaspiмагазине"],
    "seller_name": ["названиевсистемепродавца"],
    "sku": ["артикул"],
    "warehouse": ["складпередачикд", "складпередачикурьерскойдоставки"],
}

def map_headers(df: pd.DataFrame) -> Dict[str, str]:
    """Return dict canonical_key -> actual df column name (best-effort)."""
    colmap = {}
    cols_norm = {norm(c): c for c in df.columns}
    for k, variants in CANON.items():
        for v in variants:
            v_norm = norm(v)
            if v_norm in cols_norm:
                colmap[k] = cols_norm[v_norm]
                break
    return colmap

def parse_kz_date(v) -> date | None:
    if pd.isna(v):
        return None
    s = str(v).strip()
    # Kaspi exports: '21.09.2025' or datetime
    try:
        d = dtp.parse(s, dayfirst=True, yearfirst=False).date()
        return d
    except Exception:
        try:
            if isinstance(v, (pd.Timestamp, )):
                return v.date()
        except Exception:
            pass
    return None

def today_local() -> date:
    return datetime.now().date()

# ---------- Read & filter ActiveOrders ----------

def read_active_orders(orders_dir: Path) -> Tuple[pd.DataFrame, List[Path]]:
    files = sorted([p for p in orders_dir.glob("*.xlsx") if p.is_file()])
    frames = []
    for p in files:
        try:
            df = pd.read_excel(p, engine="openpyxl")
            df["__source_file__"] = p.name
            frames.append(df)
        except Exception as e:
            print(f'WARN: cannot read "{p.name}": {e}')
    if not frames:
        raise SystemExit(json.dumps({"error": f'No .xlsx files in {orders_dir}'}))
    return pd.concat(frames, ignore_index=True), files

def filter_for_shipping(df: pd.DataFrame,
                        status_wanted: str,
                        signature_wanted: str | None,
                        end_date: date) -> Tuple[pd.DataFrame, Dict]:
    colmap = map_headers(df)
    # Build status filter
    ok = pd.Series([True] * len(df))
    if "status" in colmap:
        ok &= (df[colmap["status"]].astype(str).str.strip() == status_wanted)
    # Optional signature filter
    if signature_wanted and "signature" in colmap:
        ok &= (df[colmap["signature"]].astype(str).str.strip() == signature_wanted)
    # Date ≤ end_date
    if "handover" in colmap:
        handover = df[colmap["handover"]].apply(parse_kz_date)
        ok &= handover.apply(lambda d: (d is not None) and (d <= end_date))
    df2 = df[ok].copy()
    stats = {
        "files_seen": len(df["__source_file__"].unique()) if "__source_file__" in df else None,
        "rows_in_files": int(len(df)),
        "rows_after_filters": int(len(df2)),
        "target_end_date": end_date.isoformat(),
    }
    return df2, stats


def _resolve_table(ws, table_name: str):
    tables = ws.tables
    if table_name in tables:
        return tables[table_name]
    if tables:
        # fallback to the first table on the sheet
        return next(iter(tables.values()))
    raise RuntimeError(f"No tables found on sheet {ws.title}")


def _table_bounds(table) -> Tuple[int, int, int, int]:
    start_ref, end_ref = table.ref.split(':')
    start_col_letters, start_row = coordinate_from_string(start_ref)
    end_col_letters, end_row = coordinate_from_string(end_ref)
    start_col = column_index_from_string(start_col_letters)
    end_col = column_index_from_string(end_col_letters)
    return start_col, start_row, end_col, end_row


def collect_existing_order_ids(crm_xlsx: Path,
                               sheet_name: str,
                               table_name: str,
                               order_col_abs: int) -> set[str]:
    wb = load_workbook(filename=str(crm_xlsx), read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        table = _resolve_table(ws, table_name)
        start_col, start_row, end_col, end_row = _table_bounds(table)
        # order_col_abs is sheet-level index; ensure it falls within table bounds
        if not (start_col <= order_col_abs <= end_col):
            return set()
        order_ids = set()
        for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row,
                                 min_col=order_col_abs, max_col=order_col_abs):
            cell = row[0]
            val = cell.value
            if val in (None, ""):
                continue
            if isinstance(val, str) and val.startswith("="):
                continue
            order_ids.add(str(val).strip())
        return order_ids
    finally:
        wb.close()


def validate_stage(stage_block: List[List],
                   df_filt: pd.DataFrame,
                   slice_headers: List[str],
                   existing_order_ids: set[str]) -> Dict[str, List]:
    header_norm = [norm(h) for h in slice_headers]
    # Determine relevant column indexes
    order_idx = next((i for i, h in enumerate(header_norm)
                      if h in {"orderid", "номерзаказа", "заказ", "№заказа", "заказа"}), None)
    warehouse_idx = next((i for i, h in enumerate(header_norm)
                          if h in {"складпередачикд", "складпередачикурьерскойдоставки"}), None)

    duplicates: List[str] = []
    missing_warehouse: List[Dict[str, str | None]] = []
    stage_order_ids_per_row: List[str | None] = []

    for row in stage_block:
        oid_str = None
        if order_idx is not None:
            oid = row[order_idx]
            if oid not in (None, ""):
                oid_str = str(oid).strip()
                if oid_str in existing_order_ids:
                    duplicates.append(oid_str)
        stage_order_ids_per_row.append(oid_str)

    if warehouse_idx is not None:
        for row, oid in zip(stage_block, stage_order_ids_per_row):
            wh = row[warehouse_idx]
            if wh in (None, ""):
                missing_warehouse.append({"order_id": oid})

    # Ageing based on "Плановая дата передачи курьеру"
    ageing: List[Dict] = []
    colmap = map_headers(df_filt)
    if "handover" in colmap and "order_id" in colmap:
        handover_series = df_filt[colmap["handover"]].apply(parse_kz_date)
        order_series = df_filt[colmap["order_id"]]
        today = today_local()
        for order_val, handover in zip(order_series, handover_series):
            if handover is None:
                continue
            age = (today - handover).days
            if age > 2:
                ageing.append({
                    "order_id": str(order_val),
                    "handover_date": handover.isoformat(),
                    "days_overdue": age,
                })

    return {
        "duplicate_order_ids": sorted(set(duplicates)),
        "missing_warehouse": missing_warehouse,
        "aging_orders": ageing,
    }


def archive_run(orders_dir: Path,
                source_files: List[Path],
                df_filt: pd.DataFrame) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_root = orders_dir / "archive_orders"
    archive_root.mkdir(parents=True, exist_ok=True)
    run_dir = archive_root / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)

    # Move source files into archive folder
    for src in source_files:
        dest = run_dir / src.name
        try:
            shutil.move(str(src), str(dest))
        except shutil.Error as exc:
            print(f"WARN: could not archive {src.name}: {exc}")

    # Persist a log of appended orders in CSV form
    log_cols = []
    colmap = map_headers(df_filt)
    for key in ("order_id", "handover", "warehouse", "status", "signature"):
        if key in colmap:
            log_cols.append(colmap[key])
    # Always include source file if present
    if "__source_file__" in df_filt.columns:
        log_cols.append("__source_file__")

    if log_cols:
        df_log = df_filt[log_cols].copy()
        # Rename to canonical names where possible for clarity
        rename_map = {colmap[k]: k for k in colmap if colmap[k] in log_cols}
        df_log = df_log.rename(columns=rename_map)
        if "order_id" in df_log.columns:
            df_log["order_id"] = df_log["order_id"].apply(
                lambda v: "" if pd.isna(v) else (str(int(v)) if isinstance(v, (int, float)) and float(v).is_integer() else str(v))
            )
        if "handover" in df_log.columns:
            def _fmt_handover(v):
                if pd.isna(v):
                    return ""
                parsed = parse_kz_date(v)
                return parsed.isoformat() if parsed else str(v)

            df_log["handover"] = df_log["handover"].apply(_fmt_handover)
        log_path = run_dir / "appended_orders.csv"
        df_log.to_csv(log_path, index=False)

    return run_dir

# ---------- Inspect CRM sheet to find columns we need ----------

def inspect_crm_sheet(crm_xlsx: Path, sheet_name: str, table_name: str
                      ) -> Tuple[int, int, int, List[str]]:
    """
    Return:
      date_abs_col     -> absolute column index of 'Date' header in the sheet
      slice_start_col  -> absolute column index of 'OrderID/№ заказа' header
      slice_end_col    -> absolute column index of 'Склад передачи КД' header
      slice_headers    -> list of headers (text as in Excel) from start..end inclusive
    """
    wb = load_workbook(filename=str(crm_xlsx), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" not found in {crm_xlsx}')
    ws = wb[sheet_name]
    # Header row = first row of the ListObject (table). In read-only mode we assume it's row 1.
    header_vals = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Normalize headers for search
    idx_date = None
    idx_start = None
    idx_end = None
    for i, h in enumerate(header_vals, start=1):
        hnorm = norm(h)
        if idx_date is None and hnorm in {"date", "дата"}:
            idx_date = i
        if idx_start is None and hnorm in {"orderid", "номерзаказа", "заказ", "№заказа"}:
            idx_start = i
        if idx_end is None and hnorm in {"складпередачикд", "складпередачикурьерскойдоставки"}:
            idx_end = i
    if idx_date is None or idx_start is None or idx_end is None or idx_end < idx_start:
        raise SystemExit(
            "Could not locate CRM headers. Need 'Date', 'OrderID/№ заказа', and 'Склад передачи КД' "
            f"(sheet {sheet_name}). Headers seen: {header_vals}"
        )
    slice_headers = [header_vals[j-1] for j in range(idx_start, idx_end+1)]
    wb.close()
    return idx_date, idx_start, idx_end, slice_headers

# ---------- Build the staging matrix that matches the CRM slice ----------

def build_staging(df_filt: pd.DataFrame,
                  crm_slice_headers: List[str]) -> List[List]:
    """
    Returns a 2D list of values with width = len(crm_slice_headers)
    in the exact order of the CRM slice [OrderID .. Склад передачи КД].
    We map export columns to these headers best-effort, anything missing stays blank.
    """
    colmap = map_headers(df_filt)
    # Create a convenient dict of Series by canonical key
    S: Dict[str, pd.Series] = {}
    for k, real in colmap.items():
        S[k] = df_filt[real].astype(object)

    # Also build direct lookup by normalized column header
    cols_norm_map: Dict[str, pd.Series] = {
        norm(col): df_filt[col].astype(object) for col in df_filt.columns
    }

    # Row count
    n = len(df_filt)

    # Utility: get a column by header intention
    def col_for(header_text: str) -> pd.Series:
        h = norm(header_text)
        # Try direct mapped keys
        if h in {"orderid", "номерзаказа", "заказ", "№заказа", "заказа"} and "order_id" in S:
            col = S["order_id"].copy()
            # Coerce to string to preserve leading zeros and avoid .0 suffix
            col = col.apply(lambda v: "" if pd.isna(v) else (str(int(v)) if isinstance(v, (int, float)) and float(v).is_integer() else str(v)))
            return col
        if h in {"названиевсистемепродавца"} and "seller_name" in S:
            return S["seller_name"]
        if h in {"названиетоваравkaspiмагазине"} and "offer_name" in S:
            return S["offer_name"]
        if h in {"артикул"} and "sku" in S:
            return S["sku"]
        if h in {"складпередачикд", "складпередачикурьерскойдоставки"} and "warehouse" in S:
            return S["warehouse"]
        # Direct match against the export header (most CRM columns share the same text)
        if h in cols_norm_map:
            return cols_norm_map[h]
        # Otherwise: empty
        return pd.Series([""] * n, index=df_filt.index, dtype=object)

    cols = [col_for(h) for h in crm_slice_headers]

    # Ensure OrderID looks like text numbers (Excel table will format as text/number as you prefer)
    # Leave other fields as-is.
    stage = []
    for i in range(len(df_filt)):
        row = []
        for s in cols:
            v = s.iloc[i]
            # Normalize NaN to empty
            if pd.isna(v):
                v = ""
            row.append(v)
        stage.append(row)
    return stage

# ---------- Excel append (via xlwings; invisible) ----------

def excel_append(out_wb: Path,
                 sheet_name: str,
                 table_name: str,
                 date_col_abs: int,
                 start_col_abs: int,
                 end_col_abs: int,
                 stage_block: List[List],
                 set_date: date,
                 slice_headers: List[str]) -> None:
    n = len(stage_block)
    if n == 0:
        return

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        wb = app.books.open(str(out_wb))
        sh = wb.sheets[sheet_name]

        # Locate the table via xlwings' Table API (works on both Windows & macOS)
        try:
            tbl = sh.tables[table_name]
        except KeyError:
            tables = list(sh.tables)
            if not tables:
                raise RuntimeError(f"No table found on sheet {sheet_name}")
            tbl = tables[0]

        # Determine where new rows should begin. tbl.range covers header+data.
        total_rows_before = tbl.range.rows.count
        header_row = tbl.range.row
        data_rows_before = total_rows_before - 1  # exclude header
        top_row = header_row + data_rows_before + 1
        bottom_row = top_row + n - 1

        # Writing directly below the current table causes Excel to extend it and
        # keep formulas/formatting consistent.
        date_vals = [[set_date] for _ in range(n)]
        date_range = sh.range((top_row, date_col_abs), (bottom_row, date_col_abs))
        date_range.value = date_vals
        date_range.number_format = "dd.mm.yyyy"

        width = len(slice_headers)
        for offset in range(width):
            header = slice_headers[offset]
            col_values = [row[offset] for row in stage_block]
            # Preserve formula-driven columns (OrderID column keeps its table formula)
            if norm(header) in {"orderid"}:
                continue
            # Skip writing entirely empty columns to avoid clearing formulas
            if all((v is None) or (isinstance(v, str) and v == "") for v in col_values):
                continue
            target = sh.range((top_row, start_col_abs + offset), (bottom_row, start_col_abs + offset))
            # Ensure xlwings sees a 2D list
            target.value = [[v] for v in col_values]

        wb.save()
        wb.close()
    finally:
        app.quit()

# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser(description="Append Kaspi ActiveOrders exports into CRM workbook")
    ap.add_argument(
        "--orders-dir",
        default=str(DEFAULT_ACTIVE_ORDERS_DIR),
        help="Directory containing ActiveOrders*.xlsx exports (default: %(default)s; override with KASPI_ACTIVE_ORDERS_DIR)",
    )
    ap.add_argument(
        "--out-wb",
        default=str(DEFAULT_CRM_WORKBOOK),
        help="CRM workbook path (default: %(default)s; override with KASPI_CRM_WORKBOOK)",
    )
    ap.add_argument("--sheet", default=DEFAULT_CRM_SHEET, help="CRM sheet name (default: %(default)s; override with KASPI_CRM_SHEET)")
    ap.add_argument("--table", default=DEFAULT_CRM_TABLE, help="CRM table name (default: %(default)s; override with KASPI_CRM_TABLE)")
    ap.add_argument("--date-end", default="today",
                    help='Upper bound for "Плановая дата передачи курьеру". Use "today" or YYYY-MM-DD.')
    ap.add_argument(
        "--append-date",
        default="today",
        help='Date to stamp into the CRM "Date" column. Use "today" or YYYY-MM-DD.',
    )
    ap.add_argument("--status", default=DEFAULT_KASPI_STATUS,
                    help="Status filter to include (default: %(default)s; override with KASPI_STATUS)")
    ap.add_argument("--signature", default=DEFAULT_KASPI_SIGNATURE,
                    help="Signature requirement filter (default: %(default)s; override with KASPI_SIGNATURE)")
    ap.add_argument("--dry-run", action="store_true", help="Run filters and validation only; skip Excel append")
    args = ap.parse_args()

    end_date = today_local() if args.date_end.strip().lower() == "today" else dtp.parse(args.date_end).date()
    append_date = today_local() if args.append_date.strip().lower() == "today" else dtp.parse(args.append_date).date()

    orders_dir = Path(args.orders_dir).expanduser()
    out_wb = Path(args.out_wb).expanduser()

    df_all, source_files = read_active_orders(orders_dir)
    df_filt, stats = filter_for_shipping(df_all, args.status, args.signature, end_date)

    # Inspect the CRM to locate the slice
    date_abs, start_abs, end_abs, slice_headers = inspect_crm_sheet(out_wb, args.sheet, args.table)

    existing_order_ids = collect_existing_order_ids(out_wb, args.sheet, args.table, start_abs)

    # Build stage matrix for the CRM slice
    stage = build_staging(df_filt, slice_headers)

    validation = validate_stage(stage, df_filt, slice_headers, existing_order_ids)

    payload = {
        **stats,
        "rows_to_append": len(stage),
        "target_end_date": end_date.isoformat(),
        "validation": validation,
        "files_to_archive": [p.name for p in source_files],
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    if args.dry_run:
        return

    # Append via Excel (invisible)
    excel_append(out_wb, args.sheet, args.table, date_abs, start_abs, end_abs, stage, append_date, slice_headers)
    archive_path = archive_run(orders_dir, source_files, df_filt)

    print("✅ Appended to CRM table (Excel-safe, no UI).")
    print(f"📦 Archived source exports to {archive_path}")

if __name__ == "__main__":
    main()
